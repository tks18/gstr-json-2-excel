import glob
from os import error, mkdir, startfile
from pathlib import Path
from tkinter import messagebox
from openpyxl import Workbook

# Other Imports
from ui import gst_utils_ui
from helpers import (
    get_user_json_directory,
    get_json_sales_data,
    write_basic_data_sheet,
    generate_invoices_list,
    write_invoices_to_excel,
    make_archive,
)


def write_basic_data(path_to_json):
    global work_book, basic_data, app_mode, create_excel_dir_modes

    basic_data_heading_map = {
        "gstin": "GSTIN",
        "fp": "Month",
    }
    not_required_keys = ["b2b", "cdn", "b2ba", "cdnur","b2csa","cdnra","expa"]
    basic_data = get_json_sales_data(path_to_json)
    for key in not_required_keys:
        if key in basic_data:
            basic_data.pop(key)

    if app_mode in create_excel_dir_modes:
        basic_data_sheet = work_book["Sheet"]
        basic_data_sheet.title = "Basic Data"

        write_basic_data_sheet(
            work_sheet=basic_data_sheet,
            basic_data=basic_data,
            heading_map=basic_data_heading_map,
        )


def write_b2b_invoices(path_to_json, destination):
    global work_book, file_name, create_file_dir_modes, app_mode, create_excel_dir_modes

    invoice_list = []
    b2b_heading_map = {
        "inum": "Invoice Number",
        "pos": "Place of Sale",
        "chksum": "Check Sum",
        "itms_0_itm_det_iamt": "IGST",
        "inv_typ": "Invoice Type",
        "itms_0_itm_det_samt": "SGST",
        "itms_0_itm_det_txval": "Taxable Value",
        "ctin": "GSTIN",
        "cfs3b": "GSTR 3B",
        "rchrg": "Reverse Charge",
        "itms_0_num": "Rate Number",
        "val": "Invoice Value",
        "itms_0_itm_det_camt": "CGST",
        "flprdr1": "GSTR 1 Filing Period",
        "itms_0_itm_det_csamt": "Cess",
        "fldtr1": "GSTR 1 Filing Date",
        "itms_0_itm_det_rt": "Rate",
        "idt": "Invoice Date",
        "cfs": "CFS",
        "dtcancel": "Date of Cancel(if)",
    }
    b2b_heading_list = [heading for heading in b2b_heading_map]

    sales_data = get_json_sales_data(path_to_json)
    if "b2b" in sales_data:
        invoice_list = generate_invoices_list(
            sales_data=sales_data,
            sales_type="b2b",
            invoice_term="inv",
            app_mode=app_mode,
            create_file_dir_modes=create_file_dir_modes,
            file_name=destination + "/" + file_name + "_GSTR_2A_b2b_sales.json",
        )

        if app_mode in create_excel_dir_modes:
            b2b_sheet = work_book.create_sheet("B2B")
            write_invoices_to_excel(
                work_sheet=b2b_sheet,
                invoice_list=invoice_list,
                heading_map=b2b_heading_map,
                heading_list=b2b_heading_list,
            )


def write_b2b_credit_note_invoices(path_to_json, destination):
    global work_book, file_name, create_file_dir_modes, app_mode, create_excel_dir_modes

    invoice_list = []
    b2b_credit_notes_headings_map = {
        "inum": "Invoice Number",
        "nt_dt": "Credit Note Date",
        "itms_0_itm_det_camt": "CGST",
        "chksum": "Check Sum",
        "fldtr1": "Filing Date",
        "nt_num": "Credit Note Number",
        "val": "Invoice Value",
        "itms_0_num": "Rate Number",
        "idt": "Invoice Date",
        "ntty": "Credit Note Type",
        "itms_0_itm_det_iamt": "IGST",
        "itms_0_itm_det_samt": "SGST",
        "cfs3b": "GSTR 3B",
        "flprdr1": "Filing Period",
        "p_gst": "P GST",
        "itms_0_itm_det_txval": "Taxable Value",
        "cfs": "CFS",
        "itms_0_itm_det_rt": "Rate",
        "ctin": "GSTIN",
    }

    b2b_credit_notes_headings_list = [
        heading for heading in b2b_credit_notes_headings_map
    ]

    sales_data = get_json_sales_data(path_to_json)
    if "cdn" in sales_data:
        invoice_list = generate_invoices_list(
            sales_data=sales_data,
            sales_type="cdn",
            invoice_term="nt",
            app_mode=app_mode,
            create_file_dir_modes=create_file_dir_modes,
            file_name=destination + "/" + file_name + "_GSTR_2A_b2b_sales_returns.json",
        )

        if app_mode in create_excel_dir_modes:
            b2b_credit_notes_sheet = work_book.create_sheet(title="CDNR")
            write_invoices_to_excel(
                work_sheet=b2b_credit_notes_sheet,
                invoice_list=invoice_list,
                heading_map=b2b_credit_notes_headings_map,
                heading_list=b2b_credit_notes_headings_list,
            )


def write_b2ba_invoices(path_to_json, destination):
    global work_book, file_name, create_file_dir_modes, app_mode, create_excel_dir_modes

    invoice_list = []
    b2ba_heading_map = {
        "itms_0_itm_det_csamt": "Cess",
        "idt": "Invoice Date",
        "itms_0_itm_det_camt": "CGST",
        "chksum": "Check Sum",
        "inv_typ": "Invoice Type",
        "oinum": "Original Invoice Number",
        "fldtr1": "GSTR 1 Filing Date",
        "rchrg": "Reverse Charge",
        "itms_0_num": "Rate Number",
        "itms_0_itm_det_txval": "Taxable Value",
        "cfs3b": "GSTR 3B",
        "itms_0_itm_det_iamt": "IGST",
        "ctin": "GSTIN",
        "val": "Invoice Value",
        "inum": "Invoice Number",
        "aspd": "Original Invoice Period",
        "itms_0_itm_det_rt": "Rate",
        "cfs": "CFS",
        "flprdr1": "GSTR 1 Filing Period",
        "oidt": "Old Invoice Date",
        "pos": "Place of Sale",
        "atyp": "Original Invoice Type",
        "itms_0_itm_det_samt": "SGST",
    }

    b2ba_heading_list = [heading for heading in b2ba_heading_map]
    sales_data = get_json_sales_data(path_to_json)
    if "b2ba" in sales_data:

        invoice_list = generate_invoices_list(
            sales_data=sales_data,
            sales_type="b2ba",
            invoice_term="inv",
            app_mode=app_mode,
            create_file_dir_modes=create_file_dir_modes,
            file_name=destination + "/" + file_name + "_GSTR_2A_b2ba_sales.json",
        )

        if app_mode in create_excel_dir_modes:
            b2ba_sheet = work_book.create_sheet(title="B2BA")
            write_invoices_to_excel(
                work_sheet=b2ba_sheet,
                invoice_list=invoice_list,
                heading_map=b2ba_heading_map,
                heading_list=b2ba_heading_list,
            )


def write_all_invoices(
    app_mode,
    create_excel_dir_modes,
    path_to_json,
    file_directory,
    file_name,
    zip_file_name,
):
    global work_book

    write_b2b_invoices(
        path_to_json=path_to_json,
        destination=file_directory,
    )
    write_b2b_credit_note_invoices(
        path_to_json=path_to_json,
        destination=file_directory,
    )
    write_b2ba_invoices(path_to_json=path_to_json, destination=file_directory)
    if app_mode in create_excel_dir_modes:
        work_book.save(file_directory + "/GSTR_2A_" + file_name + ".xlsx")

    if app_mode == "zipped":
        try:
            make_archive(path_to_files=file_directory, file_name=zip_file_name)
        except error:
            print(error)


def start_gstr_2_process():
    global basic_data, app_mode, app_generation_mode
    global create_file_dir_modes, create_excel_dir_modes
    global work_book, file_name, file_directory

    user_input_dirs = get_user_json_directory(
        app_generation_mode=app_generation_mode,
        source_dir_label=gstr_2_ui.source_dir_label,
        final_dir_label=gstr_2_ui.final_dir_label,
        status_text=gstr_2_ui.app_status_text,
    )

    if user_input_dirs["ready_to_process"]:
        user_confirmation_dialog = messagebox.askyesno(
            title="Confirm Proceed",
            message="Are You Sure with the Details you have Provided ?",
        )

        if user_confirmation_dialog:
            json_files = []
            if app_generation_mode == "single":
                json_files.append(user_input_dirs["source_dir"])
            elif app_generation_mode == "directory":
                gstr_2_ui.app_status_text.config(
                    text="Status - Batch Processing the Files"
                )
                for json_file in glob.glob(user_input_dirs["source_dir"]):
                    json_files.append(Path(json_file.lower()))

            for json_file in json_files:

                work_book = Workbook()
                work_book.active

                write_basic_data(path_to_json=json_file)

                file_name = basic_data["gstin"] + "_" + basic_data["fp"]
                file_directory = (
                    user_input_dirs["final_dir"] + "/" + file_name
                    if app_mode in create_file_dir_modes
                    else user_input_dirs["final_dir"]
                )

                try:
                    if app_mode in create_file_dir_modes:
                        corrected_dest_file = Path(
                            user_input_dirs["final_dir"] + "/" + file_name
                        )
                        mkdir(corrected_dest_file)
                except OSError:
                    gstr_2_ui.app_status_text.config(text="Status - Folder Error - ")
                else:
                    write_all_invoices(
                        app_mode=app_mode,
                        create_excel_dir_modes=create_excel_dir_modes,
                        path_to_json=json_file,
                        file_directory=file_directory,
                        file_name=file_name,
                        zip_file_name=Path(
                            user_input_dirs["final_dir"] + "/GSTR_2A_" + file_name
                        ),
                    )

                    if (
                        app_mode in ["json", "excel-json"]
                        and app_generation_mode == "single"
                    ):
                        startfile(
                            user_input_dirs["final_dir"] + "/" + file_name, "open"
                        )
                    else:
                        startfile(user_input_dirs["final_dir"], "open")

            restart_app_input = messagebox.askyesno(
                title="Restart App",
                message="App has Finished Processing the Files\n\nDo you like to Restart the App ?",
            )
            if restart_app_input:
                gstr_2_ui.source_dir_label.config(text=" ")
                gstr_2_ui.final_dir_label.config(text=" ")
            else:
                gstr_2_ui.close_window()

        else:
            gstr_2_ui.app_status_text.config(text="Status - Ready to Process")
            gstr_2_ui.source_dir_label.config(text=" ")
            gstr_2_ui.final_dir_label.config(text=" ")


def set_app_generation_mode():
    global app_generation_mode

    app_generation_mode = gstr_2_ui.app_generation_mode_var.get()


def set_app_processing_mode():
    global app_mode

    app_mode = gstr_2_ui.app_processing_mode_var.get()


def initiate_force_close():
    global gstr_2_ui, force_close

    gstr_2_ui.close_window()
    force_close = True


def start_window_app():
    global gstr_2_ui, force_close

    gstr_2_ui = gst_utils_ui(
        window_title="GSTR 2A Utils",
        title="GSTR 2A Utility",
        commands={
            "app_generation": set_app_generation_mode,
            "app_processing": set_app_processing_mode,
        },
        start_button=start_gstr_2_process,
    )

    gstr_2_ui.ui.protocol("WM_DELETE_WINDOW", initiate_force_close)

    gstr_2_ui.initialize_engine()

    return force_close


gstr_2_ui = None

basic_data = {}
force_close = False

app_mode = "excel"
app_generation_mode = "single"

create_file_dir_modes = ["excel-json", "zipped", "json"]
create_excel_dir_modes = ["excel-json", "zipped", "excel"]

work_book = None

file_name = ""
file_directory = ""