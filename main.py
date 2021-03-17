import pyfiglet
from itertools import cycle
from time import sleep

import json
from flatten_json import flatten

import shutil

from os import system, name, error, mkdir

from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def clear():
    if name == "nt":
        _ = system("cls")
    else:
        _ = system("clear")


def get_user_json_directory():
    global app_mode

    source_directory = input("\n\nEnter the Sales Json Directory:  ").lower()
    final_directory = input("\nEnter the Detination Directory:  ").lower()

    app_mode = input(
        "\nEnter a Mode for application (excel or zipped or excel-json):  "
    ).lower()

    app_mode_validation = False
    while not app_mode_validation:
        if app_mode not in ["excel", "zipped", "excel-json"]:
            print(
                "\nPlease Enter a Valid Mode - only 'excel' or 'zipped' or 'excel-json' are allowed\n"
            )
            app_mode = input(
                "Enter a Mode for application (excel or zipped):  "
            ).lower()
        else:
            app_mode_validation = True

    corrected_source_dir = Path(source_directory)
    return {"source_dir": corrected_source_dir, "final_dir": final_directory}


def get_json_sales_data(path_to_json):
    with open(path_to_json, mode="r") as json_data:
        sales_data = json.load(json_data)
        return sales_data


def write_basic_data(path_to_json):
    global basic_data

    basic_data_heading_map = {
        "gstin": "GSTIN",
        "fp": "Month",
        "filing_typ": "Filing Mode",
        "gt": "GT",
        "cur_gt": "Current GT",
        "fil_dt": "Filing Date",
    }
    basic_data_sheet = work_book["Sheet"]
    basic_data_sheet.title = "Basic Data"
    not_required_keys = ["b2b", "cdnr", "exp", "b2cs", "b2cl", "b2ba", "hsn"]
    basic_data = get_json_sales_data(path_to_json)
    for key in not_required_keys:
        try:
            basic_data.pop(key)
        except KeyError:
            print("> " + key + " Not Found")

    basic_data_sheet[f"{get_column_letter(4)}3"] = "Particulars"
    basic_data_sheet[f"{get_column_letter(4)}3"].font = Font(bold=True)
    basic_data_sheet[f"{get_column_letter(5)}3"] = "Value"
    basic_data_sheet[f"{get_column_letter(5)}3"].font = Font(bold=True)

    start_column = 4
    basic_data_column = 4
    basic_data_row = 4
    for (detail_heading, detail_value) in basic_data.items():
        for (basic_head, basic_value) in basic_data_heading_map.items():
            if basic_head == detail_heading:
                basic_data_sheet[
                    f"{get_column_letter(basic_data_column)}{basic_data_row}"
                ] = basic_value
                basic_data_sheet[
                    f"{get_column_letter(basic_data_column)}{basic_data_row}"
                ].font = Font(bold=True)
        basic_data_column += 1
        basic_data_sheet[
            f"{get_column_letter(basic_data_column)}{basic_data_row}"
        ] = detail_value
        basic_data_sheet[
            f"{get_column_letter(basic_data_column)}{basic_data_row}"
        ].font = Font(bold=True)
        basic_data_row += 1
        basic_data_column = start_column


def write_b2b_invoices(path_to_json, destination):
    global create_file_dir_modes

    invoice_list = []
    b2b_headings_ref_map = {}
    b2b_heading_map = {
        "chksum": "Check Sum",
        "itms_0_itm_det_iamt": "IGST",
        "rchrg": "Reverse Charge",
        "itms_0_itm_det_csamt": "Cess",
        "idt": "Invoice Date",
        "ctin": "GSTIN",
        "inv_typ": "Invoice Type",
        "itms_0_itm_det_txval": "Taxable Value",
        "cflag": "C-Flag",
        "itms_0_itm_det_camt": "CGST",
        "itms_0_itm_det_rt": "Rate",
        "updby": "Updated by",
        "cfs": "CFS",
        "flag": "Flag",
        "inum": "Invoice No.",
        "itms_0_itm_det_samt": "SGST",
        "pos": "Place of Sale",
        "itms_0_num": "Rate Number",
        "val": "Invoice Value",
    }
    b2b_heading_list = [heading for heading in b2b_heading_map]

    sales_data = get_json_sales_data(path_to_json)
    if "b2b" in sales_data:
        b2b_sheet = work_book.create_sheet(title="B2B")
        for sales in sales_data["b2b"]:
            current_supplier = sales.copy()
            current_supplier.pop("inv")
            for invoice in sales["inv"]:
                # sample_test = {
                #     "no": invoice["inum"],
                #     "value": invoice["itms"][0]["itm_det"]["txval"],
                # }
                flattened_inv = flatten(invoice)
                invoice_list.append({**current_supplier, **flattened_inv})

        if app_mode in create_file_dir_modes:
            with open(destination + "/b2b_sales.json", mode="w") as b2b_sales_data:
                json.dump(invoice_list, b2b_sales_data)

        heading_column = 1
        heading_row = 1
        for headings in set().union(*(d.keys() for d in invoice_list)):
            if headings in b2b_heading_list:
                for (formatted_keys, formatted_vals) in b2b_heading_map.items():
                    if formatted_keys == headings:
                        b2b_headings_ref_map.update(
                            {headings: f"{get_column_letter(heading_column)}"}
                        )
                        b2b_sheet[
                            f"{get_column_letter(heading_column)}{heading_row}"
                        ] = formatted_vals
                        b2b_sheet[
                            f"{get_column_letter(heading_column)}{heading_row}"
                        ].font = Font(bold=True)
            else:
                b2b_headings_ref_map.update(
                    {headings: f"{get_column_letter(heading_column)}"}
                )
                b2b_sheet[
                    f"{get_column_letter(heading_column)}{heading_row}"
                ] = headings
                b2b_sheet[
                    f"{get_column_letter(heading_column)}{heading_row}"
                ].font = Font(bold=True)
            heading_column += 1

        invoice_column = 1
        invoice_row = 2
        for invoice in invoice_list:
            for (item, value) in invoice.items():
                for (headings, excel_ref) in b2b_headings_ref_map.items():
                    if headings == item:
                        b2b_sheet[f"{excel_ref}{invoice_row}"] = value
                invoice_column += 1
            invoice_column = 1
            invoice_row += 1

    else:
        print("> No B2B Invoices Found in JSON")


def write_b2b_credit_note_invoices(path_to_json, destination):
    global create_file_dir_modes

    invoice_list = []
    b2b_credit_notes_headings_ref_map = {}
    b2b_credit_notes_headings_map = {
        "itms_0_itm_det_iamt": "IGST",
        "updby": "Updated by",
        "cfs": "CFS",
        "nt_num": "Invoice No",
        "itms_0_itm_det_samt": "SGST",
        "nt_dt": "Return Date",
        "idt": "Invoice Date",
        "itms_0_itm_det_rt": "Rate",
        "val": "Invoice Value",
        "flag": "Flag",
        "p_gst": "P GST",
        "itms_0_itm_det_camt": "CGST",
        "ntty": "Credit Type",
        "ctin": "GSTIN",
        "itms_0_num": "Rate Number",
        "cflag": "C Flag",
        "itms_0_itm_det_txval": "Taxable Value",
        "inum": "Invoice Number",
        "chksum": "Check Sum",
    }

    b2b_credit_notes_headings_list = [
        heading for heading in b2b_credit_notes_headings_map
    ]

    sales_data = get_json_sales_data(path_to_json)
    if "cdnr" in sales_data:
        b2b_credit_notes_sheet = work_book.create_sheet(title="CDNR")

        for sales_returns in sales_data["cdnr"]:
            current_supplier = sales_returns.copy()
            current_supplier.pop("nt")
            for invoice in sales_returns["nt"]:
                flattened_invoice = flatten(invoice)
                invoice_list.append({**current_supplier, **flattened_invoice})

        if app_mode in create_file_dir_modes:
            with open(
                file=destination + "/b2b_sales_returns.json", mode="w"
            ) as b2b_sales_returns_data:
                json.dump(obj=invoice_list, fp=b2b_sales_returns_data)

        heading_column = 1
        heading_row = 1
        for headings in set().union(*(d.keys() for d in invoice_list)):
            if headings in b2b_credit_notes_headings_list:
                for (
                    formatted_keys,
                    formatted_vals,
                ) in b2b_credit_notes_headings_map.items():
                    if formatted_keys == headings:
                        b2b_credit_notes_headings_ref_map.update(
                            {headings: f"{get_column_letter(heading_column)}"}
                        )
                        b2b_credit_notes_sheet[
                            f"{get_column_letter(heading_column)}{heading_row}"
                        ] = formatted_vals
                        b2b_credit_notes_sheet[
                            f"{get_column_letter(heading_column)}{heading_row}"
                        ].font = Font(bold=True)
            else:
                b2b_credit_notes_headings_ref_map.update(
                    {headings: f"{get_column_letter(heading_column)}"}
                )
                b2b_credit_notes_sheet[
                    f"{get_column_letter(heading_column)}{heading_row}"
                ] = headings
                b2b_credit_notes_sheet[
                    f"{get_column_letter(heading_column)}{heading_row}"
                ].font = Font(bold=True)
            heading_column += 1

        invoice_column = 1
        invoice_row = 2
        for invoice in invoice_list:
            for (item, value) in invoice.items():
                for (headings, excel_ref) in b2b_credit_notes_headings_ref_map.items():
                    if headings == item:
                        b2b_credit_notes_sheet[f"{excel_ref}{invoice_row}"] = value
                invoice_column += 1
            invoice_column = 1
            invoice_row += 1
    else:
        print("> No B2B Credit Notes Found in JSON")


def write_b2cs_invoices(path_to_json, destination):
    global create_file_dir_modes

    invoice_list = []
    b2cs_sales_heading_ref_map = {}

    b2cs_sales_heading_map = {
        "samt": "SGST",
        "camt": "CGST",
        "typ": "Type",
        "flag": "Flag",
        "sply_ty": "Supply Type",
        "chksum": "Check Sum",
        "iamt": "IGST",
        "txval": "Taxable Value",
        "rt": "Rate",
        "pos": "Place of Supply",
    }

    b2cs_sales_heading_list = [heading for heading in b2cs_sales_heading_map]

    sales_data = get_json_sales_data(path_to_json)
    if "b2cs" in sales_data:
        b2cs_sales_sheet = work_book.create_sheet(title="B2CS")

        for invoice in sales_data["b2cs"]:
            flattened_invoice = flatten(invoice)
            invoice_list.append(flattened_invoice)

        if app_mode in create_file_dir_modes:
            with open(
                file=destination + "/b2cs_sales.json", mode="w"
            ) as b2cs_sales_data:
                json.dump(obj=invoice_list, fp=b2cs_sales_data)

        heading_column = 1
        heading_row = 1
        for headings in set().union(*(d.keys() for d in invoice_list)):
            if headings in b2cs_sales_heading_list:
                for (
                    formatted_keys,
                    formatted_vals,
                ) in b2cs_sales_heading_map.items():
                    if formatted_keys == headings:
                        b2cs_sales_heading_ref_map.update(
                            {headings: f"{get_column_letter(heading_column)}"}
                        )
                        b2cs_sales_sheet[
                            f"{get_column_letter(heading_column)}{heading_row}"
                        ] = formatted_vals
                        b2cs_sales_sheet[
                            f"{get_column_letter(heading_column)}{heading_row}"
                        ].font = Font(bold=True)
            else:
                b2cs_sales_heading_ref_map.update(
                    {headings: f"{get_column_letter(heading_column)}"}
                )
                b2cs_sales_sheet[
                    f"{get_column_letter(heading_column)}{heading_row}"
                ] = headings
                b2cs_sales_sheet[
                    f"{get_column_letter(heading_column)}{heading_row}"
                ].font = Font(bold=True)
            heading_column += 1

        invoice_column = 1
        invoice_row = 2
        for invoice in invoice_list:
            for (item, value) in invoice.items():
                for (headings, excel_ref) in b2cs_sales_heading_ref_map.items():
                    if headings == item:
                        b2cs_sales_sheet[f"{excel_ref}{invoice_row}"] = value
                invoice_column += 1
            invoice_column = 1
            invoice_row += 1
    else:
        print("> No B2CS Invoices Found in JSON")


def write_export_invoices(path_to_json, destination):
    global create_file_dir_modes

    invoice_list = []
    export_heading_ref_map = {}
    export_headings_map = {
        "itms_0_csamt": "Cess",
        "itms_0_iamt": "IGST",
        "val": "Invoice Value",
        "itms_0_rt": "Rate",
        "idt": "Invoice Date",
        "exp_typ": "Exports Type",
        "flag": "Flag",
        "chksum": "Check Sum",
        "itms_0_txval": "Taxable Value",
        "inum": "Invoice Number",
    }
    export_headings_list = [heading for heading in export_headings_map]

    sales_data = get_json_sales_data(path_to_json)
    if "exp" in sales_data:
        export_sheet = work_book.create_sheet(title="Exports")

        for export_sales in sales_data["exp"]:
            current_supplier = export_sales.copy()
            current_supplier.pop("inv")
            for invoice in export_sales["inv"]:
                flattened_invoice = flatten(invoice)
                invoice_list.append({**current_supplier, **flattened_invoice})

        if app_mode in create_file_dir_modes:
            with open(
                Path(destination + "/export_sales.json"), mode="w"
            ) as export_sales_data:
                json.dump(obj=invoice_list, fp=export_sales_data)

        heading_column = 1
        heading_row = 1
        for headings in set().union(*(d.keys() for d in invoice_list)):
            if headings in export_headings_list:
                for (
                    formatted_keys,
                    formatted_vals,
                ) in export_headings_map.items():
                    if formatted_keys == headings:
                        export_heading_ref_map.update(
                            {headings: f"{get_column_letter(heading_column)}"}
                        )
                        export_sheet[
                            f"{get_column_letter(heading_column)}{heading_row}"
                        ] = formatted_vals
                        export_sheet[
                            f"{get_column_letter(heading_column)}{heading_row}"
                        ].font = Font(bold=True)
            else:
                export_heading_ref_map.update(
                    {headings: f"{get_column_letter(heading_column)}"}
                )
                export_sheet[
                    f"{get_column_letter(heading_column)}{heading_row}"
                ] = headings
                export_sheet[
                    f"{get_column_letter(heading_column)}{heading_row}"
                ].font = Font(bold=True)
            heading_column += 1

        invoice_column = 1
        invoice_row = 2
        for invoice in invoice_list:
            for (item, value) in invoice.items():
                for (headings, excel_ref) in export_heading_ref_map.items():
                    if headings == item:
                        export_sheet[f"{excel_ref}{invoice_row}"] = value
                invoice_column += 1
            invoice_column = 1
            invoice_row += 1
    else:
        print("> No Export Invoices Found in JSON")


def make_archive(path_to_files, file_name):
    shutil.make_archive(base_name=file_name, format="zip", root_dir=path_to_files)
    shutil.rmtree(path_to_files)


clear()

title = "GSTR 1 Json to Excel Utility"
subtitle = pyfiglet.figlet_format("Shan.tk")

print("\n" + title + "\n" + "by" + "\n" + subtitle + "\n")

spinner_time = 1
for frame in cycle(r"-\|/"):
    print("\r", frame, sep="", end=" Initializing Application", flush=True)
    sleep(0.2)
    if spinner_time == 15:
        break
    spinner_time += 1

restart_app = True

while restart_app:

    work_book = Workbook()
    work_book.active

    basic_data = {}
    app_mode = ""

    user_input_dirs = get_user_json_directory()
    write_basic_data(path_to_json=user_input_dirs["source_dir"])

    create_file_dir_modes = ["excel-json", "zipped"]
    file_name = basic_data["gstin"] + "_" + basic_data["fp"]
    file_directory = (
        user_input_dirs["final_dir"] + "/" + file_name
        if app_mode in create_file_dir_modes
        else user_input_dirs["final_dir"]
    )

    try:
        if app_mode in create_file_dir_modes:
            corrected_dest_file = Path(user_input_dirs["final_dir"] + "/" + file_name)
            mkdir(corrected_dest_file)
    except OSError as folder_error:
        print(folder_error)
    else:
        write_b2b_invoices(
            path_to_json=user_input_dirs["source_dir"],
            destination=file_directory,
        )
        write_b2b_credit_note_invoices(
            path_to_json=user_input_dirs["source_dir"],
            destination=file_directory,
        )
        write_b2cs_invoices(
            path_to_json=user_input_dirs["source_dir"],
            destination=file_directory,
        )
        write_export_invoices(
            path_to_json=user_input_dirs["source_dir"],
            destination=file_directory,
        )
        work_book.save(file_directory + "/" + file_name + ".xlsx")

        if app_mode == "zipped":
            try:
                make_archive(path_to_files=file_directory, file_name=file_name)
            except error:
                print(error)

        restart_app_input = input("\nDo You want to Generate another? (y/n)  ").lower()
        restart_app_validation = False
        while not restart_app_validation:
            if restart_app_input not in ["y", "n"]:
                print("\nOnly (y/n) is Accepted\n")
                restart_app_input = input(
                    "\nDo You want to Generate another? (y/n)  "
                ).lower()
            else:
                restart_app_validation = True

        restart_app = True if restart_app_input == "y" else False
        if restart_app:
            clear()