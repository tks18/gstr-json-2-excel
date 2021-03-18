import pyfiglet
from itertools import cycle
from time import sleep

import glob

import json
from flatten_json import flatten

import shutil

from os import system, name, error, mkdir

from pathlib import Path

import tkinter as tk
from tkinter import Label, StringVar, filedialog

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def clear():
    if name == "nt":
        _ = system("cls")
    else:
        _ = system("clear")


def init_app():
    clear()

    title = "GSTR 1 Json to Excel Utility"
    subtitle = pyfiglet.figlet_format("Shan.tk")

    print("\n" + title + "\n" + "by" + "\n" + subtitle + "\n")

    spinner_time = 1
    for frame in cycle(r"-\|/"):
        print("\r", frame, sep="", end=" Initializing Application", flush=True)
        sleep(0.2)
        if spinner_time == 15:
            break
        spinner_time += 1


def get_user_json_directory():
    global app_mode, allowed_modes, app_generation_mode, allowed_generation_modes

    app_generation_mode = input(
        "\n\nEnter a App Working Mode (single or directory):  "
    ).lower()

    app_generation_validation = False
    while not app_generation_validation:
        if app_generation_mode not in allowed_generation_modes:
            print(
                "\nPlease Enter a Valid Working Mode - only 'single' or 'directory' is allowed\n"
            )
            app_generation_mode = input(
                "\n\nEnter a App Working Mode (single or directory):  "
            ).lower()
        else:
            app_generation_validation = True

    if app_generation_mode == "single":
        source_directory = filedialog.askopenfilename()
    elif app_generation_mode == "directory":
        print("\n!! Ensure the Directory only has JSON files !!")
        source_directory = filedialog.askdirectory()

    final_directory = input("\nEnter the Detination Directory:  ").lower()

    corrected_source_dir = (
        Path(source_directory.lower())
        if app_generation_mode == "single"
        else source_directory.lower() + "\\*.json"
    )
    return {"source_dir": corrected_source_dir, "final_dir": final_directory}


def get_json_sales_data(path_to_json):
    with open(file=path_to_json, mode="r") as json_data:
        sales_data = json.load(json_data)
        return sales_data


def write_basic_data(path_to_json):
    global work_book, basic_data, app_mode, create_excel_dir_modes

    basic_data_heading_map = {
        "gstin": "GSTIN",
        "fp": "Month",
        "filing_typ": "Filing Mode",
        "gt": "GT",
        "cur_gt": "Current GT",
        "fil_dt": "Filing Date",
    }
    not_required_keys = ["b2b", "cdnr", "exp", "b2cs", "b2cl", "b2ba", "hsn"]
    basic_data = get_json_sales_data(path_to_json)
    for key in not_required_keys:
        try:
            basic_data.pop(key)
        except KeyError:
            print(
                "> "
                + basic_data["gstin"]
                + "_"
                + basic_data["fp"]
                + " "
                + key
                + " Not Found"
            )

    if app_mode in create_excel_dir_modes:
        basic_data_sheet = work_book["Sheet"]
        basic_data_sheet.title = "Basic Data"

        basic_data_sheet[f"{get_column_letter(4)}3"] = "Particulars"
        basic_data_sheet[f"{get_column_letter(4)}3"].font = Font(bold=True)
        basic_data_sheet[f"{get_column_letter(5)}3"] = "Value"
        basic_data_sheet[f"{get_column_letter(5)}3"].font = Font(bold=True)

        start_column = 4
        basic_data_column = 4
        basic_data_row = 4
        for (detail_heading, detail_value) in basic_data.items():
            for (basic_head, basic_value) in basic_data_heading_map.items():
                if basic_head == detail_heading:
                    basic_data_sheet[
                        f"{get_column_letter(basic_data_column)}{basic_data_row}"
                    ] = basic_value
                    basic_data_sheet[
                        f"{get_column_letter(basic_data_column)}{basic_data_row}"
                    ].font = Font(bold=True)
            basic_data_column += 1
            basic_data_sheet[
                f"{get_column_letter(basic_data_column)}{basic_data_row}"
            ] = detail_value
            basic_data_sheet[
                f"{get_column_letter(basic_data_column)}{basic_data_row}"
            ].font = Font(bold=True)
            basic_data_row += 1
            basic_data_column = start_column


def write_b2b_invoices(path_to_json, destination):
    global work_book, file_name, create_file_dir_modes, app_mode, create_excel_dir_modes

    invoice_list = []
    b2b_headings_ref_map = {}
    b2b_heading_map = {
        "chksum": "Check Sum",
        "itms_0_itm_det_iamt": "IGST",
        "rchrg": "Reverse Charge",
        "itms_0_itm_det_csamt": "Cess",
        "idt": "Invoice Date",
        "ctin": "GSTIN",
        "inv_typ": "Invoice Type",
        "itms_0_itm_det_txval": "Taxable Value",
        "cflag": "C-Flag",
        "itms_0_itm_det_camt": "CGST",
        "itms_0_itm_det_rt": "Rate",
        "updby": "Updated by",
        "cfs": "CFS",
        "flag": "Flag",
        "inum": "Invoice No.",
        "itms_0_itm_det_samt": "SGST",
        "pos": "Place of Sale",
        "itms_0_num": "Rate Number",
        "val": "Invoice Value",
    }
    b2b_heading_list = [heading for heading in b2b_heading_map]

    sales_data = get_json_sales_data(path_to_json)
    if "b2b" in sales_data:
        for sales in sales_data["b2b"]:
            current_supplier = sales.copy()
            current_supplier.pop("inv")
            for invoice in sales["inv"]:
                current_invoice = invoice.copy()
                if len(invoice["itms"]) > 1:
                    current_invoice.pop("itms")
                    for current_inv_item in invoice["itms"]:
                        new_invoice = current_invoice.copy()
                        new_invoice["itms"] = [current_inv_item]
                        flattened_inv = flatten(new_invoice)
                        invoice_list.append({**current_supplier, **flattened_inv})
                else:
                    flattened_inv = flatten(invoice)
                    invoice_list.append({**current_supplier, **flattened_inv})

        if app_mode in create_file_dir_modes:
            with open(
                destination + "/" + file_name + "_b2b_sales.json", mode="w"
            ) as b2b_sales_data:
                json.dump(invoice_list, b2b_sales_data)

        if app_mode in create_excel_dir_modes:
            b2b_sheet = work_book.create_sheet("B2B")
            heading_column = 1
            heading_row = 1
            for headings in set().union(*(d.keys() for d in invoice_list)):
                if headings in b2b_heading_list:
                    for (formatted_keys, formatted_vals) in b2b_heading_map.items():
                        if formatted_keys == headings:
                            b2b_headings_ref_map.update(
                                {headings: f"{get_column_letter(heading_column)}"}
                            )
                            b2b_sheet[
                                f"{get_column_letter(heading_column)}{heading_row}"
                            ] = formatted_vals
                            b2b_sheet[
                                f"{get_column_letter(heading_column)}{heading_row}"
                            ].font = Font(bold=True)
                else:
                    b2b_headings_ref_map.update(
                        {headings: f"{get_column_letter(heading_column)}"}
                    )
                    b2b_sheet[
                        f"{get_column_letter(heading_column)}{heading_row}"
                    ] = headings
                    b2b_sheet[
                        f"{get_column_letter(heading_column)}{heading_row}"
                    ].font = Font(bold=True)
                heading_column += 1

            invoice_column = 1
            invoice_row = 2
            for invoice in invoice_list:
                for (item, value) in invoice.items():
                    for (headings, excel_ref) in b2b_headings_ref_map.items():
                        if headings == item:
                            b2b_sheet[f"{excel_ref}{invoice_row}"] = value
                    invoice_column += 1
                invoice_column = 1
                invoice_row += 1

    else:
        print("> " + file_name + " No B2B Invoices Found in JSON")


def write_b2b_credit_note_invoices(path_to_json, destination):
    global work_book, file_name, create_file_dir_modes, app_mode, create_excel_dir_modes

    invoice_list = []
    b2b_credit_notes_headings_ref_map = {}
    b2b_credit_notes_headings_map = {
        "itms_0_itm_det_iamt": "IGST",
        "updby": "Updated by",
        "cfs": "CFS",
        "nt_num": "Invoice No",
        "itms_0_itm_det_samt": "SGST",
        "nt_dt": "Return Date",
        "idt": "Invoice Date",
        "itms_0_itm_det_rt": "Rate",
        "val": "Invoice Value",
        "flag": "Flag",
        "p_gst": "P GST",
        "itms_0_itm_det_camt": "CGST",
        "ntty": "Credit Type",
        "ctin": "GSTIN",
        "itms_0_num": "Rate Number",
        "cflag": "C Flag",
        "itms_0_itm_det_txval": "Taxable Value",
        "inum": "Invoice Number",
        "chksum": "Check Sum",
    }

    b2b_credit_notes_headings_list = [
        heading for heading in b2b_credit_notes_headings_map
    ]

    sales_data = get_json_sales_data(path_to_json)
    if "cdnr" in sales_data:

        for sales_returns in sales_data["cdnr"]:
            current_supplier = sales_returns.copy()
            current_supplier.pop("nt")
            for invoice in sales_returns["nt"]:
                current_invoice = invoice.copy()
                if len(invoice["itms"]) > 1:
                    current_invoice.pop("itms")
                    for current_inv_item in invoice["itms"]:
                        new_invoice = current_invoice.copy()
                        new_invoice["itms"] = [current_inv_item]
                        flattened_inv = flatten(new_invoice)
                        invoice_list.append({**current_supplier, **flattened_inv})
                else:
                    flattened_inv = flatten(invoice)
                    invoice_list.append({**current_supplier, **flattened_inv})

        if app_mode in create_file_dir_modes:
            with open(
                file=destination + "/" + file_name + "_b2b_sales_returns.json", mode="w"
            ) as b2b_sales_returns_data:
                json.dump(obj=invoice_list, fp=b2b_sales_returns_data)

        if app_mode in create_excel_dir_modes:
            b2b_credit_notes_sheet = work_book.create_sheet(title="CDNR")
            heading_column = 1
            heading_row = 1
            for headings in set().union(*(d.keys() for d in invoice_list)):
                if headings in b2b_credit_notes_headings_list:
                    for (
                        formatted_keys,
                        formatted_vals,
                    ) in b2b_credit_notes_headings_map.items():
                        if formatted_keys == headings:
                            b2b_credit_notes_headings_ref_map.update(
                                {headings: f"{get_column_letter(heading_column)}"}
                            )
                            b2b_credit_notes_sheet[
                                f"{get_column_letter(heading_column)}{heading_row}"
                            ] = formatted_vals
                            b2b_credit_notes_sheet[
                                f"{get_column_letter(heading_column)}{heading_row}"
                            ].font = Font(bold=True)
                else:
                    b2b_credit_notes_headings_ref_map.update(
                        {headings: f"{get_column_letter(heading_column)}"}
                    )
                    b2b_credit_notes_sheet[
                        f"{get_column_letter(heading_column)}{heading_row}"
                    ] = headings
                    b2b_credit_notes_sheet[
                        f"{get_column_letter(heading_column)}{heading_row}"
                    ].font = Font(bold=True)
                heading_column += 1

            invoice_column = 1
            invoice_row = 2
            for invoice in invoice_list:
                for (item, value) in invoice.items():
                    for (
                        headings,
                        excel_ref,
                    ) in b2b_credit_notes_headings_ref_map.items():
                        if headings == item:
                            b2b_credit_notes_sheet[f"{excel_ref}{invoice_row}"] = value
                    invoice_column += 1
                invoice_column = 1
                invoice_row += 1
    else:
        print("> " + file_name + " No B2B Credit Notes Found in JSON")


def write_b2cs_invoices(path_to_json, destination):
    global work_book, file_name, create_file_dir_modes, app_mode, create_excel_dir_modes

    invoice_list = []
    b2cs_sales_heading_ref_map = {}

    b2cs_sales_heading_map = {
        "samt": "SGST",
        "camt": "CGST",
        "typ": "Type",
        "flag": "Flag",
        "sply_ty": "Supply Type",
        "chksum": "Check Sum",
        "iamt": "IGST",
        "txval": "Taxable Value",
        "rt": "Rate",
        "pos": "Place of Supply",
    }

    b2cs_sales_heading_list = [heading for heading in b2cs_sales_heading_map]

    sales_data = get_json_sales_data(path_to_json)
    if "b2cs" in sales_data:

        for invoice in sales_data["b2cs"]:
            flattened_invoice = flatten(invoice)
            invoice_list.append(flattened_invoice)

        if app_mode in create_file_dir_modes:
            with open(
                file=destination + "/" + file_name + "_b2cs_sales.json", mode="w"
            ) as b2cs_sales_data:
                json.dump(obj=invoice_list, fp=b2cs_sales_data)

        if app_mode in create_excel_dir_modes:
            b2cs_sales_sheet = work_book.create_sheet(title="B2CS")
            heading_column = 1
            heading_row = 1
            for headings in set().union(*(d.keys() for d in invoice_list)):
                if headings in b2cs_sales_heading_list:
                    for (
                        formatted_keys,
                        formatted_vals,
                    ) in b2cs_sales_heading_map.items():
                        if formatted_keys == headings:
                            b2cs_sales_heading_ref_map.update(
                                {headings: f"{get_column_letter(heading_column)}"}
                            )
                            b2cs_sales_sheet[
                                f"{get_column_letter(heading_column)}{heading_row}"
                            ] = formatted_vals
                            b2cs_sales_sheet[
                                f"{get_column_letter(heading_column)}{heading_row}"
                            ].font = Font(bold=True)
                else:
                    b2cs_sales_heading_ref_map.update(
                        {headings: f"{get_column_letter(heading_column)}"}
                    )
                    b2cs_sales_sheet[
                        f"{get_column_letter(heading_column)}{heading_row}"
                    ] = headings
                    b2cs_sales_sheet[
                        f"{get_column_letter(heading_column)}{heading_row}"
                    ].font = Font(bold=True)
                heading_column += 1

            invoice_column = 1
            invoice_row = 2
            for invoice in invoice_list:
                for (item, value) in invoice.items():
                    for (headings, excel_ref) in b2cs_sales_heading_ref_map.items():
                        if headings == item:
                            b2cs_sales_sheet[f"{excel_ref}{invoice_row}"] = value
                    invoice_column += 1
                invoice_column = 1
                invoice_row += 1
    else:
        print("> " + file_name + " No B2CS Invoices Found in JSON")


def write_export_invoices(path_to_json, destination):
    global work_book, file_name, create_file_dir_modes, app_mode, create_excel_dir_modes

    invoice_list = []
    export_heading_ref_map = {}
    export_headings_map = {
        "itms_0_csamt": "Cess",
        "itms_0_iamt": "IGST",
        "val": "Invoice Value",
        "itms_0_rt": "Rate",
        "idt": "Invoice Date",
        "exp_typ": "Exports Type",
        "flag": "Flag",
        "chksum": "Check Sum",
        "itms_0_txval": "Taxable Value",
        "inum": "Invoice Number",
    }
    export_headings_list = [heading for heading in export_headings_map]

    sales_data = get_json_sales_data(path_to_json)
    if "exp" in sales_data:

        for export_sales in sales_data["exp"]:
            current_supplier = export_sales.copy()
            current_supplier.pop("inv")
            for invoice in export_sales["inv"]:
                current_invoice = invoice.copy()
                if len(invoice["itms"]) > 1:
                    current_invoice.pop("itms")
                    for current_inv_item in invoice["itms"]:
                        new_invoice = current_invoice.copy()
                        new_invoice["itms"] = [current_inv_item]
                        flattened_inv = flatten(new_invoice)
                        invoice_list.append({**current_supplier, **flattened_inv})
                else:
                    flattened_inv = flatten(invoice)
                    invoice_list.append({**current_supplier, **flattened_inv})

        if app_mode in create_file_dir_modes:
            with open(
                Path(destination + "/" + file_name + "_export_sales.json"), mode="w"
            ) as export_sales_data:
                json.dump(obj=invoice_list, fp=export_sales_data)

        if app_mode in create_excel_dir_modes:
            export_sheet = work_book.create_sheet(title="Exports")
            heading_column = 1
            heading_row = 1
            for headings in set().union(*(d.keys() for d in invoice_list)):
                if headings in export_headings_list:
                    for (
                        formatted_keys,
                        formatted_vals,
                    ) in export_headings_map.items():
                        if formatted_keys == headings:
                            export_heading_ref_map.update(
                                {headings: f"{get_column_letter(heading_column)}"}
                            )
                            export_sheet[
                                f"{get_column_letter(heading_column)}{heading_row}"
                            ] = formatted_vals
                            export_sheet[
                                f"{get_column_letter(heading_column)}{heading_row}"
                            ].font = Font(bold=True)
                else:
                    export_heading_ref_map.update(
                        {headings: f"{get_column_letter(heading_column)}"}
                    )
                    export_sheet[
                        f"{get_column_letter(heading_column)}{heading_row}"
                    ] = headings
                    export_sheet[
                        f"{get_column_letter(heading_column)}{heading_row}"
                    ].font = Font(bold=True)
                heading_column += 1

            invoice_column = 1
            invoice_row = 2
            for invoice in invoice_list:
                for (item, value) in invoice.items():
                    for (headings, excel_ref) in export_heading_ref_map.items():
                        if headings == item:
                            export_sheet[f"{excel_ref}{invoice_row}"] = value
                    invoice_column += 1
                invoice_column = 1
                invoice_row += 1
    else:
        print("> " + file_name + " No Export Invoices Found in JSON")


def write_b2ba_invoices(path_to_json, destination):
    global work_book, file_name, create_file_dir_modes, app_mode, create_excel_dir_modes

    invoice_list = []
    b2ba_heading_ref_map = {}
    b2ba_heading_map = {
        "idt": "Invoice Date",
        "inv_typ": "Invoice Type",
        "itms_0_itm_det_rt": "Rate",
        "itms_0_itm_det_camt": "CGST",
        "pos": "Place of Sale",
        "val": "Invoice Value",
        "updby": "Updated by",
        "itms_0_itm_det_txval": "Taxable Value",
        "cflag": "C Flag",
        "itms_0_num": "Rate Number",
        "oidt": "Old Invoice Date",
        "ctin": "GSTIN",
        "cfs": "CFS",
        "rchrg": "Reverse Charge",
        "itms_0_itm_det_samt": "SGST",
        "inum": "Invoice Number",
        "oinum": "Old Invoice Number",
        "flag": "Flag",
        "chksum": "Check Sum",
    }

    b2ba_heading_list = [heading for heading in b2ba_heading_map]
    sales_data = get_json_sales_data(path_to_json)
    if "b2ba" in sales_data:

        for b2ba_sales in sales_data["b2ba"]:
            current_supplier = b2ba_sales.copy()
            current_supplier.pop("inv")
            for invoice in b2ba_sales["inv"]:
                current_invoice = invoice.copy()
                if len(invoice["itms"]) > 1:
                    current_invoice.pop("itms")
                    for current_inv_item in invoice["itms"]:
                        new_invoice = current_invoice.copy()
                        new_invoice["itms"] = [current_inv_item]
                        flattened_inv = flatten(new_invoice)
                        invoice_list.append({**current_supplier, **flattened_inv})
                else:
                    flattened_inv = flatten(invoice)
                    invoice_list.append({**current_supplier, **flattened_inv})

        if app_mode in create_file_dir_modes:
            with open(
                Path(destination + "/" + file_name + "_b2ba_sales.json"), mode="w"
            ) as export_sales_data:
                json.dump(obj=invoice_list, fp=export_sales_data)

        if app_mode in create_excel_dir_modes:
            b2ba_sheet = work_book.create_sheet(title="B2BA")
            heading_column = 1
            heading_row = 1
            for headings in set().union(*(d.keys() for d in invoice_list)):
                if headings in b2ba_heading_list:
                    for (
                        formatted_keys,
                        formatted_vals,
                    ) in b2ba_heading_map.items():
                        if formatted_keys == headings:
                            b2ba_heading_ref_map.update(
                                {headings: f"{get_column_letter(heading_column)}"}
                            )
                            b2ba_sheet[
                                f"{get_column_letter(heading_column)}{heading_row}"
                            ] = formatted_vals
                            b2ba_sheet[
                                f"{get_column_letter(heading_column)}{heading_row}"
                            ].font = Font(bold=True)
                else:
                    b2ba_heading_ref_map.update(
                        {headings: f"{get_column_letter(heading_column)}"}
                    )
                    b2ba_sheet[
                        f"{get_column_letter(heading_column)}{heading_row}"
                    ] = headings
                    b2ba_sheet[
                        f"{get_column_letter(heading_column)}{heading_row}"
                    ].font = Font(bold=True)
                heading_column += 1

            invoice_column = 1
            invoice_row = 2
            for invoice in invoice_list:
                for (item, value) in invoice.items():
                    for (headings, excel_ref) in b2ba_heading_ref_map.items():
                        if headings == item:
                            b2ba_sheet[f"{excel_ref}{invoice_row}"] = value
                    invoice_column += 1
                invoice_column = 1
                invoice_row += 1
    else:
        print("> " + file_name + " No B2BA Invoices Found in JSON")


def make_archive(path_to_files, file_name):
    shutil.make_archive(base_name=file_name, format="zip", root_dir=path_to_files)
    shutil.rmtree(path_to_files)


def write_all_invoices(
    app_mode, create_excel_dir_modes, path_to_json, file_directory, file_name
):
    global work_book

    write_b2b_invoices(
        path_to_json=path_to_json,
        destination=file_directory,
    )
    write_b2b_credit_note_invoices(
        path_to_json=path_to_json,
        destination=file_directory,
    )
    write_b2cs_invoices(
        path_to_json=path_to_json,
        destination=file_directory,
    )
    write_export_invoices(
        path_to_json=path_to_json,
        destination=file_directory,
    )
    write_b2ba_invoices(path_to_json=path_to_json, destination=file_directory)
    if app_mode in create_excel_dir_modes:
        work_book.save(file_directory + "/" + file_name + ".xlsx")

    if app_mode == "zipped":
        try:
            make_archive(path_to_files=file_directory, file_name=file_name)
        except error:
            print(error)


def start_gstr_1_process():
    global basic_data, app_mode, app_generation_mode
    global allowed_generation_modes, allowed_modes, create_file_dir_modes, create_excel_dir_modes
    global work_book, file_name, file_directory

    restart_app = True
    while restart_app:

        user_input_dirs = get_user_json_directory()

        json_files = []
        if app_generation_mode == "single":
            json_files.append(user_input_dirs["source_dir"])
        elif app_generation_mode == "directory":
            for json_file in glob.glob(user_input_dirs["source_dir"]):
                json_files.append(Path(json_file.lower()))

        print("\nRunning Application Log Drain <<>>\n")
        for json_file in json_files:

            work_book = Workbook()
            work_book.active

            write_basic_data(path_to_json=json_file)

            file_name = basic_data["gstin"] + "_" + basic_data["fp"]
            file_directory = (
                user_input_dirs["final_dir"] + "/" + file_name
                if app_mode in create_file_dir_modes
                else user_input_dirs["final_dir"]
            )

            try:
                if app_mode in create_file_dir_modes:
                    corrected_dest_file = Path(
                        user_input_dirs["final_dir"] + "/" + file_name
                    )
                    mkdir(corrected_dest_file)
            except OSError as folder_error:
                print(folder_error)
            else:
                write_all_invoices(
                    app_mode=app_mode,
                    create_excel_dir_modes=create_excel_dir_modes,
                    path_to_json=json_file,
                    file_directory=file_directory,
                    file_name=file_name,
                )

        restart_app_input = input("\nDo You want to Restart the app? (y/n)  ").lower()
        restart_app_validation = False
        while not restart_app_validation:
            if restart_app_input not in ["y", "n"]:
                print("\nOnly (y/n) is Accepted\n")
                restart_app_input = input(
                    "\nDo You want to Generate another? (y/n)  "
                ).lower()
            else:
                restart_app_validation = True

        restart_app = True if restart_app_input == "y" else False
        if restart_app:
            clear()


def set_app_generation_mode():
    global app_generation_mode, app_generation_mode_var

    print(app_generation_mode_var.get(), app_generation_mode)
    app_generation_mode = app_generation_mode_var.get()


def start_window_app():
    global app_generation_mode_var

    init_app()
    gstr_1_ui = tk.Toplevel()
    gstr_1_ui.title("Sub Window")
    gstr_1_ui.config(padx=100, pady=80)

    main_title = tk.Label(
        master=gstr_1_ui,
        text="GSTR 1 Utility",
        font=("Courier new", 23, "bold"),
        pady=40,
    )
    main_title.grid(row=0, column=0, columnspan=3)

    app_generation_mode_label = tk.Label(gstr_1_ui, text="App Working Mode")
    app_generation_mode_label.grid(row=1, column=0, columnspan=3)
    app_generation_mode_var = StringVar(gstr_1_ui, "single")
    app_generation_modes = {
        "Single": {"value": "single", "row": 2, "column": 0, "span": 2},
        "Directory": {"value": "directory", "row": 2, "column": 1, "span": 2},
    }
    for (modes, mode_vals) in app_generation_modes.items():
        tk.Radiobutton(
            gstr_1_ui,
            text=modes,
            variable=app_generation_mode_var,
            value=mode_vals["value"],
            command=set_app_generation_mode,
        ).grid(
            row=mode_vals["row"],
            column=mode_vals["column"],
            columnspan=mode_vals["span"],
        )

    # app_mode_label = Label(gstr_1_ui, text="What Type of File Generation You Need")
    start_gstr_1_process()


basic_data = {}
app_mode = ""
app_generation_mode = ""
app_generation_mode_var = None

allowed_generation_modes = ["single", "directory"]
allowed_modes = ["excel", "json", "zipped", "excel-json"]
create_file_dir_modes = ["excel-json", "zipped", "json"]
create_excel_dir_modes = ["excel-json", "zipped", "excel"]

work_book = None

file_name = ""
file_directory = ""
