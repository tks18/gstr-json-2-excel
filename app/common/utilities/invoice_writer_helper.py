from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from app.common.threader import threader


def invoice_to_excel_function(work_sheet, invoice_list, heading_map, heading_list):
    heading_ref_map = {}
    heading_column = 1
    heading_row = 1
    for headings in set().union(*(d.keys() for d in invoice_list)):
        if headings in heading_list:
            for (formatted_keys, formatted_vals) in heading_map.items():
                if formatted_keys == headings:
                    heading_ref_map.update(
                        {headings: f"{get_column_letter(heading_column)}"}
                    )
                    work_sheet[
                        f"{get_column_letter(heading_column)}{heading_row}"
                    ] = formatted_vals
                    work_sheet[
                        f"{get_column_letter(heading_column)}{heading_row}"
                    ].font = Font(bold=True)
        else:
            heading_ref_map.update({headings: f"{get_column_letter(heading_column)}"})
            work_sheet[f"{get_column_letter(heading_column)}{heading_row}"] = headings
            work_sheet[f"{get_column_letter(heading_column)}{heading_row}"].font = Font(
                bold=True
            )
        heading_column += 1

    invoice_column = 1
    invoice_row = 2
    for invoice in invoice_list:
        for (item, value) in invoice.items():
            for (headings, excel_ref) in heading_ref_map.items():
                if headings == item:
                    work_sheet[f"{excel_ref}{invoice_row}"] = value
            invoice_column += 1
        invoice_column = 1
        invoice_row += 1


def write_invoices_to_excel(work_sheet, invoice_list, heading_map, heading_list):
    invoices_to_excel_thread = threader(
        function=invoice_to_excel_function,
        work_sheet=work_sheet,
        invoice_list=invoice_list,
        heading_map=heading_map,
        heading_list=heading_list,
    )
    invoices_to_excel_thread.start()
    invoices_to_excel_thread.join()
    if invoices_to_excel_thread.error is not None:
        raise invoices_to_excel_thread.error