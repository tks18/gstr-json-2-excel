from app.helpers.threader import threader
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def basic_data_function(work_book, basic_data, heading_map):
    basic_data_sheet = work_book["Sheet"]
    basic_data_sheet.title = "Basic Data"

    basic_data_sheet[f"{get_column_letter(4)}3"] = "Particulars"
    basic_data_sheet[f"{get_column_letter(4)}3"].font = Font(bold=True)
    basic_data_sheet[f"{get_column_letter(5)}3"] = "Value"
    basic_data_sheet[f"{get_column_letter(5)}3"].font = Font(bold=True)

    start_column = 4
    basic_data_column = 4
    basic_data_row = 4
    for (detail_heading, detail_value) in basic_data.items():
        for (basic_head, basic_value) in heading_map.items():
            if basic_head == detail_heading:
                basic_data_sheet[
                    f"{get_column_letter(basic_data_column)}{basic_data_row}"
                ] = basic_value
                basic_data_sheet[
                    f"{get_column_letter(basic_data_column)}{basic_data_row}"
                ].font = Font(bold=True)
        basic_data_column += 1
        basic_data_sheet[
            f"{get_column_letter(basic_data_column)}{basic_data_row}"
        ] = detail_value
        basic_data_sheet[
            f"{get_column_letter(basic_data_column)}{basic_data_row}"
        ].font = Font(bold=True)
        basic_data_row += 1
        basic_data_column = start_column


def write_basic_data_sheet(work_book, basic_data, heading_map):
    basic_data_thread = threader(
        function=basic_data_function,
        work_book=work_book,
        basic_data=basic_data,
        heading_map=heading_map,
    )
    basic_data_thread.start()
    basic_data_thread.join()
    if basic_data_thread.error is not None:
        raise basic_data_thread.error