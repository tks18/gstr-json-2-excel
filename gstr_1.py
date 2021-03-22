import glob
from os import error, mkdir, startfile
from pathlib import Path
from tkinter import messagebox
from openpyxl import Workbook

# Other Imports
from ui import gst_utils_ui
from helpers import (
    get_user_json_directory,
    get_json_sales_data,
    write_basic_data_sheet,
    generate_invoices_list,
    write_invoices_to_excel,
    make_archive,
)


def write_basic_data(path_to_json):
    global work_book, basic_data, app_mode, create_excel_dir_modes

    basic_data_heading_map = {
        "gstin": "GSTIN",
        "fp": "Month",
        "filing_typ": "Filing Mode",
        "gt": "GT",
        "cur_gt": "Current GT",
        "fil_dt": "Filing Date",
    }
    not_required_keys = ["b2b", "cdnr", "exp", "b2cs", "b2cl", "b2ba", "hsn"]
    basic_data = get_json_sales_data(path_to_json)
    for key in not_required_keys:
        if key in basic_data:
            basic_data.pop(key)

    if app_mode in create_excel_dir_modes:
        basic_data_sheet = work_book["Sheet"]
        basic_data_sheet.title = "Basic Data"

        write_basic_data_sheet(
            work_sheet=basic_data_sheet,
            basic_data=basic_data,
            heading_map=basic_data_heading_map,
        )


def write_b2b_invoices(path_to_json, destination):
    global work_book, file_name, create_file_dir_modes, app_mode, create_excel_dir_modes

    invoice_list = []
    b2b_heading_map = {
        "chksum": "Check Sum",
        "itms_0_itm_det_iamt": "IGST",
        "rchrg": "Reverse Charge",
        "itms_0_itm_det_csamt": "Cess",
        "idt": "Invoice Date",
        "ctin": "GSTIN",
        "inv_typ": "Invoice Type",
        "itms_0_itm_det_txval": "Taxable Value",
        "cflag": "C-Flag",
        "itms_0_itm_det_camt": "CGST",
        "itms_0_itm_det_rt": "Rate",
        "updby": "Updated by",
        "cfs": "CFS",
        "flag": "Flag",
        "inum": "Invoice No.",
        "itms_0_itm_det_samt": "SGST",
        "pos": "Place of Sale",
        "itms_0_num": "Rate Number",
        "val": "Invoice Value",
    }
    b2b_heading_list = [heading for heading in b2b_heading_map]

    sales_data = get_json_sales_data(path_to_json)
    if "b2b" in sales_data:
        invoice_list = generate_invoices_list(
            sales_data=sales_data,
            sales_type="b2b",
            invoice_term="inv",
            app_mode=app_mode,
            create_file_dir_modes=create_file_dir_modes,
            file_name=destination + "/" + file_name + "GSTR1_b2b_sales.json",
        )

        if app_mode in create_excel_dir_modes:
            b2b_sheet = work_book.create_sheet("B2B")
            write_invoices_to_excel(
                work_sheet=b2b_sheet,
                invoice_list=invoice_list,
                heading_map=b2b_heading_map,
                heading_list=b2b_heading_list,
            )


def write_b2b_credit_note_invoices(path_to_json, destination):
    global work_book, file_name, create_file_dir_modes, app_mode, create_excel_dir_modes

    invoice_list = []
    b2b_credit_notes_headings_map = {
        "itms_0_itm_det_iamt": "IGST",
        "updby": "Updated by",
        "cfs": "CFS",
        "nt_num": "Credit Note No",
        "itms_0_itm_det_samt": "SGST",
        "nt_dt": "Credit Note Date",
        "idt": "Invoice Date",
        "itms_0_itm_det_rt": "Rate",
        "val": "Invoice Value",
        "flag": "Flag",
        "p_gst": "P GST",
        "itms_0_itm_det_camt": "CGST",
        "ntty": "Credit Note Type",
        "ctin": "GSTIN",
        "itms_0_num": "Rate Number",
        "cflag": "C Flag",
        "itms_0_itm_det_txval": "Taxable Value",
        "inum": "Original Invoice Number",
        "chksum": "Check Sum",
    }

    b2b_credit_notes_headings_list = [
        heading for heading in b2b_credit_notes_headings_map
    ]

    sales_data = get_json_sales_data(path_to_json)
    if "cdnr" in sales_data:
        invoice_list = generate_invoices_list(
            sales_data=sales_data,
            sales_type="cdnr",
            invoice_term="nt",
            app_mode=app_mode,
            create_file_dir_modes=create_file_dir_modes,
            file_name=destination + "/" + file_name + "GSTR1_b2b_sales_returns.json",
        )

        if app_mode in create_excel_dir_modes:
            b2b_credit_notes_sheet = work_book.create_sheet(title="CDNR")
            write_invoices_to_excel(
                work_sheet=b2b_credit_notes_sheet,
                invoice_list=invoice_list,
                heading_map=b2b_credit_notes_headings_map,
                heading_list=b2b_credit_notes_headings_list,
            )


def write_b2cs_invoices(path_to_json, destination):
    global work_book, file_name, create_file_dir_modes, app_mode, create_excel_dir_modes

    invoice_list = []
    b2cs_sales_heading_map = {
        "samt": "SGST",
        "camt": "CGST",
        "typ": "Type",
        "flag": "Flag",
        "sply_ty": "Supply Type",
        "chksum": "Check Sum",
        "iamt": "IGST",
        "txval": "Taxable Value",
        "rt": "Rate",
        "pos": "Place of Supply",
    }

    b2cs_sales_heading_list = [heading for heading in b2cs_sales_heading_map]

    sales_data = get_json_sales_data(path_to_json)
    if "b2cs" in sales_data:

        invoice_list = generate_invoices_list(
            sales_data=sales_data,
            sales_type="b2cs",
            invoice_term="inv",
            app_mode=app_mode,
            create_file_dir_modes=create_file_dir_modes,
            file_name=destination + "/" + file_name + "GSTR1_b2cs_sales.json",
        )

        if app_mode in create_excel_dir_modes:
            b2cs_sales_sheet = work_book.create_sheet(title="B2CS")
            write_invoices_to_excel(
                work_sheet=b2cs_sales_sheet,
                invoice_list=invoice_list,
                heading_map=b2cs_sales_heading_map,
                heading_list=b2cs_sales_heading_list,
            )


def write_export_invoices(path_to_json, destination):
    global work_book, file_name, create_file_dir_modes, app_mode, create_excel_dir_modes

    invoice_list = []
    export_headings_map = {
        "itms_0_csamt": "Cess",
        "itms_0_iamt": "IGST",
        "val": "Invoice Value",
        "itms_0_rt": "Rate",
        "idt": "Invoice Date",
        "exp_typ": "Exports Type",
        "flag": "Flag",
        "chksum": "Check Sum",
        "itms_0_txval": "Taxable Value",
        "inum": "Invoice Number",
    }
    export_headings_list = [heading for heading in export_headings_map]

    sales_data = get_json_sales_data(path_to_json)
    if "exp" in sales_data:

        invoice_list = generate_invoices_list(
            sales_data=sales_data,
            sales_type="exp",
            invoice_term="inv",
            app_mode=app_mode,
            create_file_dir_modes=create_file_dir_modes,
            file_name=destination + "/" + file_name + "GSTR1_export_sales.json",
        )

        if app_mode in create_excel_dir_modes:
            export_sheet = work_book.create_sheet(title="Exports")
            write_invoices_to_excel(
                work_sheet=export_sheet,
                invoice_list=invoice_list,
                heading_map=export_headings_map,
                heading_list=export_headings_list,
            )


def write_b2ba_invoices(path_to_json, destination):
    global work_book, file_name, create_file_dir_modes, app_mode, create_excel_dir_modes

    invoice_list = []
    b2ba_heading_map = {
        "idt": "Invoice Date",
        "inv_typ": "Invoice Type",
        "itms_0_itm_det_rt": "Rate",
        "itms_0_itm_det_camt": "CGST",
        "pos": "Place of Sale",
        "val": "Invoice Value",
        "updby": "Updated by",
        "itms_0_itm_det_txval": "Taxable Value",
        "cflag": "C Flag",
        "itms_0_num": "Rate Number",
        "oidt": "Old Invoice Date",
        "ctin": "GSTIN",
        "cfs": "CFS",
        "rchrg": "Reverse Charge",
        "itms_0_itm_det_samt": "SGST",
        "inum": "Invoice Number",
        "oinum": "Old Invoice Number",
        "flag": "Flag",
        "chksum": "Check Sum",
    }

    b2ba_heading_list = [heading for heading in b2ba_heading_map]
    sales_data = get_json_sales_data(path_to_json)
    if "b2ba" in sales_data:

        invoice_list = generate_invoices_list(
            sales_data=sales_data,
            sales_type="b2ba",
            invoice_term="inv",
            app_mode=app_mode,
            create_file_dir_modes=create_file_dir_modes,
            file_name=destination + "/" + file_name + "GSTR1_b2ba_sales.json",
        )

        if app_mode in create_excel_dir_modes:
            b2ba_sheet = work_book.create_sheet(title="B2BA")
            write_invoices_to_excel(
                work_sheet=b2ba_sheet,
                invoice_list=invoice_list,
                heading_map=b2ba_heading_map,
                heading_list=b2ba_heading_list,
            )


def write_all_invoices(
    app_mode,
    create_excel_dir_modes,
    path_to_json,
    file_directory,
    file_name,
    zip_file_name,
):
    global work_book

    write_b2b_invoices(
        path_to_json=path_to_json,
        destination=file_directory,
    )
    write_b2b_credit_note_invoices(
        path_to_json=path_to_json,
        destination=file_directory,
    )
    write_b2cs_invoices(
        path_to_json=path_to_json,
        destination=file_directory,
    )
    write_export_invoices(
        path_to_json=path_to_json,
        destination=file_directory,
    )
    write_b2ba_invoices(path_to_json=path_to_json, destination=file_directory)
    if app_mode in create_excel_dir_modes:
        work_book.save(file_directory + "/" + file_name + ".xlsx")

    if app_mode == "zipped":
        try:
            make_archive(path_to_files=file_directory, file_name=zip_file_name)
        except error:
            print(error)


def start_gstr_1_process():
    global basic_data, app_mode, app_generation_mode
    global create_file_dir_modes, create_excel_dir_modes
    global work_book, file_name, file_directory

    user_input_dirs = get_user_json_directory(
        app_generation_mode=app_generation_mode,
        source_dir_label=gstr_1_ui.source_dir_label,
        final_dir_label=gstr_1_ui.final_dir_label,
        status_text=gstr_1_ui.app_status_text,
    )

    if user_input_dirs["ready_to_process"]:
        user_confirmation_dialog = messagebox.askyesno(
            title="Confirm Proceed",
            message="Are You Sure with the Details you have Provided ?",
        )

        if user_confirmation_dialog:
            json_files = []
            if app_generation_mode == "single":
                json_files.append(user_input_dirs["source_dir"])
            elif app_generation_mode == "directory":
                gstr_1_ui.app_status_text.config(
                    text="Status - Batch Processing the Files"
                )
                for json_file in glob.glob(user_input_dirs["source_dir"]):
                    json_files.append(Path(json_file.lower()))

            for json_file in json_files:

                work_book = Workbook()
                work_book.active

                write_basic_data(path_to_json=json_file)

                file_name = basic_data["gstin"] + "_" + basic_data["fp"]
                file_directory = (
                    user_input_dirs["final_dir"] + "/" + file_name
                    if app_mode in create_file_dir_modes
                    else user_input_dirs["final_dir"]
                )

                try:
                    if app_mode in create_file_dir_modes:
                        corrected_dest_file = Path(
                            user_input_dirs["final_dir"] + "/" + file_name
                        )
                        mkdir(corrected_dest_file)
                except OSError:
                    gstr_1_ui.app_status_text.config(text="Status - Folder Error - ")
                else:
                    write_all_invoices(
                        app_mode=app_mode,
                        create_excel_dir_modes=create_excel_dir_modes,
                        path_to_json=json_file,
                        file_directory=file_directory,
                        file_name=file_name,
                        zip_file_name=Path(
                            user_input_dirs["final_dir"] + "/GSTR1_" + file_name
                        ),
                    )

                    if (
                        app_mode in ["json", "excel-json"]
                        and app_generation_mode == "single"
                    ):
                        startfile(
                            user_input_dirs["final_dir"] + "/" + file_name, "open"
                        )
                    else:
                        startfile(user_input_dirs["final_dir"], "open")

            restart_app_input = messagebox.askyesno(
                title="Restart App",
                message="App has Finished Processing the Files\n\nDo you like to Restart the App ?",
            )
            if restart_app_input:
                gstr_1_ui.source_dir_label.config(text=" ")
                gstr_1_ui.final_dir_label.config(text=" ")
            else:
                gstr_1_ui.close_window()

        else:
            gstr_1_ui.app_status_text.config(text="Status - Ready to Process")
            gstr_1_ui.source_dir_label.config(text=" ")
            gstr_1_ui.final_dir_label.config(text=" ")


def set_app_generation_mode():
    global app_generation_mode

    app_generation_mode = gstr_1_ui.app_generation_mode_var.get()


def set_app_processing_mode():
    global app_mode

    app_mode = gstr_1_ui.app_processing_mode_var.get()


def initiate_force_close():
    global gstr_1_ui, force_close

    gstr_1_ui.close_window()
    force_close = True


def start_window_app():
    global gstr_1_ui, force_close

    gstr_1_ui = gst_utils_ui(
        window_title="GSTR 1 Utils",
        title="GSTR 1 Utility",
        commands={
            "app_generation": set_app_generation_mode,
            "app_processing": set_app_processing_mode,
        },
        start_button=start_gstr_1_process,
    )

    gstr_1_ui.ui.protocol("WM_DELETE_WINDOW", initiate_force_close)

    gstr_1_ui.initialize_engine()

    return force_close


gstr_1_ui = None

basic_data = {}
force_close = False

app_mode = "excel"
app_generation_mode = "single"

create_file_dir_modes = ["excel-json", "zipped", "json"]
create_excel_dir_modes = ["excel-json", "zipped", "excel"]

work_book = None

file_name = ""
file_directory = ""