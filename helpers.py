import json
import shutil
from flatten_json import flatten
from tkinter import filedialog
from pathlib import Path
from os import path as osPath
import sys

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = osPath.abspath(".")

    return osPath.join(base_path, relative_path)


def get_json_sales_data(path_to_json):
    with open(file=path_to_json, mode="r") as json_data:
        sales_data = json.load(json_data)
        return sales_data


def get_user_json_directory(
    app_generation_mode, source_dir_label, final_dir_label, status_text
):

    if len(app_generation_mode) > 1:
        status_text.config(text="Status - Getting Directories")

        if app_generation_mode == "single":
            source_directory = filedialog.askopenfilename(
                title="Select the JSON File You want to Process",
                filetypes=[("Json", "*.json")],
            )
        elif app_generation_mode == "directory":
            source_directory = filedialog.askdirectory(
                title="Select the Folder where all the JSON Files are Located",
            )

        if len(source_directory) > 2:
            source_dir_label.config(text="Source:  " + source_directory.lower())

            final_directory = filedialog.askdirectory(
                title="Select the Destination Directory to Store the Excel Files after Processing"
            )

            if len(final_directory) > 2:
                final_dir_label.config(text="Destination: " + final_directory.lower())

            corrected_source_dir = (
                Path(source_directory.lower())
                if app_generation_mode == "single"
                else source_directory.lower() + "\\*.json"
            )

            ready_to_process = (
                True
                if len(source_directory) > 2 and len(final_directory) > 2
                else False
            )

            return {
                "ready_to_process": ready_to_process,
                "source_dir": corrected_source_dir,
                "final_dir": final_directory.lower(),
            }

        else:
            status_text.config(text="Status - Ready to Process")
            return {"ready_to_process": False, "source_dir": "", "final_dir": ""}


def write_basic_data_sheet(work_sheet, basic_data, heading_map):
    work_sheet[f"{get_column_letter(4)}3"] = "Particulars"
    work_sheet[f"{get_column_letter(4)}3"].font = Font(bold=True)
    work_sheet[f"{get_column_letter(5)}3"] = "Value"
    work_sheet[f"{get_column_letter(5)}3"].font = Font(bold=True)

    start_column = 4
    basic_data_column = 4
    basic_data_row = 4
    for (detail_heading, detail_value) in basic_data.items():
        for (basic_head, basic_value) in heading_map.items():
            if basic_head == detail_heading:
                work_sheet[
                    f"{get_column_letter(basic_data_column)}{basic_data_row}"
                ] = basic_value
                work_sheet[
                    f"{get_column_letter(basic_data_column)}{basic_data_row}"
                ].font = Font(bold=True)
        basic_data_column += 1
        work_sheet[
            f"{get_column_letter(basic_data_column)}{basic_data_row}"
        ] = detail_value
        work_sheet[
            f"{get_column_letter(basic_data_column)}{basic_data_row}"
        ].font = Font(bold=True)
        basic_data_row += 1
        basic_data_column = start_column


def generate_invoices_list(
    sales_data, sales_type, invoice_term, app_mode, create_file_dir_modes, file_name
):
    invoice_list = []
    if not sales_type == "b2cs":
        for sales in sales_data[sales_type]:
            current_supplier = sales.copy()
            current_supplier.pop(invoice_term)
            for invoice in sales[invoice_term]:
                current_invoice = invoice.copy()
                if len(invoice["itms"]) > 1:
                    current_invoice.pop("itms")
                    for current_inv_item in invoice["itms"]:
                        new_invoice = current_invoice.copy()
                        new_invoice["itms"] = [current_inv_item]
                        flattened_inv = flatten(new_invoice)
                        invoice_list.append({**current_supplier, **flattened_inv})
                else:
                    flattened_inv = flatten(invoice)
                    invoice_list.append({**current_supplier, **flattened_inv})

        if app_mode in create_file_dir_modes:
            with open(file_name, mode="w") as b2b_sales_data:
                json.dump(invoice_list, b2b_sales_data)

    else:
        for invoice in sales_data[sales_type]:
            flattened_invoice = flatten(invoice)
            invoice_list.append(flattened_invoice)

    return invoice_list


def write_invoices_to_excel(work_sheet, invoice_list, heading_map, heading_list):
    heading_ref_map = {}
    heading_column = 1
    heading_row = 1
    for headings in set().union(*(d.keys() for d in invoice_list)):
        if headings in heading_list:
            for (formatted_keys, formatted_vals) in heading_map.items():
                if formatted_keys == headings:
                    heading_ref_map.update(
                        {headings: f"{get_column_letter(heading_column)}"}
                    )
                    work_sheet[
                        f"{get_column_letter(heading_column)}{heading_row}"
                    ] = formatted_vals
                    work_sheet[
                        f"{get_column_letter(heading_column)}{heading_row}"
                    ].font = Font(bold=True)
        else:
            heading_ref_map.update({headings: f"{get_column_letter(heading_column)}"})
            work_sheet[f"{get_column_letter(heading_column)}{heading_row}"] = headings
            work_sheet[f"{get_column_letter(heading_column)}{heading_row}"].font = Font(
                bold=True
            )
        heading_column += 1

    invoice_column = 1
    invoice_row = 2
    for invoice in invoice_list:
        for (item, value) in invoice.items():
            for (headings, excel_ref) in heading_ref_map.items():
                if headings == item:
                    work_sheet[f"{excel_ref}{invoice_row}"] = value
            invoice_column += 1
        invoice_column = 1
        invoice_row += 1


def make_archive(path_to_files, file_name):
    shutil.make_archive(base_name=file_name, format="zip", root_dir=path_to_files)
    shutil.rmtree(path_to_files)